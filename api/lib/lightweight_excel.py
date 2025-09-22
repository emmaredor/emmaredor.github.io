"""
Lightweight Excel data loader that doesn't use pandas
"""

import json
import yaml
from typing import Dict, List, Optional, Tuple, Any
import re
from datetime import datetime
import openpyxl


class LightweightExcelLoader:
    """Excel loader using only openpyxl instead of pandas to reduce bundle size."""
    
    @staticmethod
    def load_students_from_excel_content(excel_content: bytes) -> List[Dict[str, Any]]:
        """
        Load student data from Excel file content.
        
        Args:
            excel_content: Raw Excel file content as bytes
            
        Returns:
            List of student dictionaries with grades
        """
        try:
            # Save to temporary file and load with openpyxl
            import tempfile
            import os
            
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                tmp_file.write(excel_content)
                tmp_file.flush()
                
                # Load workbook
                workbook = openpyxl.load_workbook(tmp_file.name)
                sheet = workbook.active
                
                # Get header row (row 2 in 1-based indexing)
                headers = []
                for cell in sheet[2]:
                    headers.append(cell.value if cell.value else "")
                
                students = []
                
                # Process each data row (starting from row 3)
                for row_idx in range(3, sheet.max_row + 1):
                    row_data = {}
                    for col_idx, header in enumerate(headers, 1):
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        row_data[header] = cell_value
                    
                    # Extract student data
                    student_data = LightweightExcelLoader._extract_student_from_row(row_data, headers)
                    if student_data:
                        students.append(student_data)
                
                # Clean up
                os.unlink(tmp_file.name)
                return students
                
        except Exception as e:
            raise ValueError(f"Error processing Excel file: {str(e)}")
    
    @staticmethod
    def _extract_student_from_row(row_data: Dict[str, Any], headers: List[str]) -> Optional[Dict[str, Any]]:
        """Extract student information and grades from a row."""
        
        # Find student info columns
        name_col = LightweightExcelLoader._find_column_by_patterns(headers, ["Etud_Nom", "nom", "name"])
        firstname_col = LightweightExcelLoader._find_column_by_patterns(headers, ["Etud_Prénom", "prénom", "firstname"])
        birth_date_col = LightweightExcelLoader._find_column_by_patterns(headers, ["Etud_Naissance", "naissance", "birth"])
        birth_city_col = LightweightExcelLoader._find_column_by_patterns(headers, ["Etud_Ville", "ville", "city"])
        
        if not name_col or not firstname_col:
            return None
        
        # Extract basic student info
        name = row_data.get(name_col)
        firstname = row_data.get(firstname_col)
        
        if not name or not firstname:
            return None
        
        # Format birth date
        birth_date = row_data.get(birth_date_col) if birth_date_col else None
        birth_city = row_data.get(birth_city_col) if birth_city_col else "Unknown"
        
        formatted_birth_date = LightweightExcelLoader._format_birth_date(birth_date)
        
        # Extract grades
        grades = LightweightExcelLoader._extract_grades_from_row(row_data, headers)
        
        if not grades:
            return None
        
        return {
            'student_info': {
                'firstname': str(firstname).strip(),
                'name': str(name).strip().upper(),
                'date_of_birth': formatted_birth_date,
                'place_of_birth': f"{birth_city} (FRANCE)" if birth_city else "Unknown (FRANCE)",
                'gender': 'Mr',  # Default
                'pronoun': 'he'  # Default
            },
            'grades': grades
        }
    
    @staticmethod
    def _find_column_by_patterns(headers: List[str], patterns: List[str]) -> Optional[str]:
        """Find column name that matches any of the given patterns."""
        for header in headers:
            if header:
                header_lower = str(header).lower()
                for pattern in patterns:
                    if pattern.lower() in header_lower:
                        return header
        return None
    
    @staticmethod
    def _format_birth_date(date_value: Any) -> str:
        """Format birth date to readable string."""
        if not date_value:
            return "Unknown"
        
        try:
            if isinstance(date_value, datetime):
                return date_value.strftime("%-d%s of %B %Y").replace("1st", "1st").replace("2nd", "2nd").replace("3rd", "3rd")
            elif isinstance(date_value, str):
                # Try to parse string date
                for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"]:
                    try:
                        dt = datetime.strptime(date_value, fmt)
                        day = dt.day
                        suffix = "th"
                        if day in [1, 21, 31]:
                            suffix = "st"
                        elif day in [2, 22]:
                            suffix = "nd"
                        elif day in [3, 23]:
                            suffix = "rd"
                        return f"{day}{suffix} of {dt.strftime('%B %Y')}"
                    except ValueError:
                        continue
            return str(date_value)
        except:
            return "Unknown"
    
    @staticmethod
    def _extract_grades_from_row(row_data: Dict[str, Any], headers: List[str]) -> Dict[str, List[float]]:
        """Extract grades from row data."""
        grades = {}
        
        # Find course columns (pattern: ObjX_Libellé, ObjX_Note_Ado/20, ObjX_Crédits)
        course_pattern = re.compile(r'Obj(\d+)_Libellé')
        
        for header in headers:
            if not header:
                continue
                
            match = course_pattern.match(str(header))
            if match:
                obj_num = match.group(1)
                
                # Get course name
                course_name = row_data.get(header)
                if not course_name or str(course_name).strip() == "":
                    continue
                
                # Look for corresponding grade and credits columns
                grade_col = f"Obj{obj_num}_Note_Ado/20"
                credits_col = f"Obj{obj_num}_Crédits"
                type_col = f"Obj{obj_num}_Type"
                
                # Check if this is an ELP course
                course_type = row_data.get(type_col)
                if course_type != "ELP":
                    continue
                
                # Get grade and credits
                grade = row_data.get(grade_col)
                credits = row_data.get(credits_col)
                
                if grade is not None and str(grade).strip() != "":
                    try:
                        grade_float = float(grade)
                        credits_int = int(credits) if credits and str(credits).strip() != "" else 6
                        
                        course_name_clean = str(course_name).strip()
                        grades[course_name_clean] = [grade_float, credits_int]
                    except (ValueError, TypeError):
                        continue
        
        return grades